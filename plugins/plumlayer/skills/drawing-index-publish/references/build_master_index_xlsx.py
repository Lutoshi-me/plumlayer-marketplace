"""Publish the Master Drawing Index Excel workbook.

Inputs:
  - Per-issue Drawing Index CSVs (positional, chronological order)
  - The Franken Set CSV (--franken)
  - Optional anomaly CSVs (--anomaly, repeatable)

Output:
  - <out>/Master Drawing Index.xlsx

Every Sheet Number cell gets a HYPERLINK to its source PDF + page.

See ../SKILL.md for the full design rationale.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import OrderedDict
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DISCIPLINE_ORDER = {
    name: i for i, name in enumerate([
        "GENERAL", "LIFE SAFETY", "CIVIL", "ARCHITECTURAL DEMO", "ARCHITECTURAL",
        "STRUCTURAL", "FIRE PROTECTION", "PLUMBING", "HVAC", "ELECTRICAL",
        "FIRE ALARM", "SECURITY", "TELECOMMUNICATIONS", "ELEVATOR",
    ])
}

# Background colors per discipline (light, readable). Cycled if more disciplines.
DISCIPLINE_FILLS = [
    "FFF6E5", "E5F6FF", "F0FFE5", "FFE5F0", "F5F5F5", "FFFFE5",
    "E5FFF0", "FFE5E5", "E5E5FF", "F0E5FF", "FFE5FF", "E5FFFF",
    "FFF0E5", "EBEBEB",
]
DISCIPLINE_COLOR_MAP = {d: DISCIPLINE_FILLS[i % len(DISCIPLINE_FILLS)]
                       for i, d in enumerate(DISCIPLINE_ORDER)}


def sheet_sort_key(sn: str) -> tuple:
    m = re.match(r"^([A-Z]+-?)([0-9]+)(?:\.([0-9]+))?([A-Z]?)$", sn)
    if not m:
        return (sn, 0, 0, "")
    prefix, intpart, decpart, suffix = m.groups()
    return (prefix, int(intpart), int(decpart) if decpart else 0, suffix or "")


def safe_filename(s: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", s).strip()


def issue_label_from(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename))[0]
    label = re.sub(r"^Drawing Index\s*-\s*", "", base).strip()
    return label


def tab_name_from_issue(label: str) -> str:
    """Excel sheet names: max 31 chars, no [ ] / \\ ? * :."""
    # Strip leading date if present
    label = re.sub(r"^\d{4}-\d{2}-\d{2}\s*", "", label)
    label = re.sub(r"[\[\]\\/\?\*:]", "-", label)
    return label[:31] or "Sheet"


def build_pdf_index(search_roots: list[str]) -> dict[str, str]:
    index: dict[str, str] = {}
    for root in search_roots:
        if not os.path.isdir(root):
            continue
        for dirpath, _, files in os.walk(root):
            for f in files:
                if f.lower().endswith(".pdf"):
                    index[f] = os.path.join(dirpath, f)
    return index


def to_file_uri(abs_path: str, page: int) -> str:
    """file:///<path>#page=N — forward slashes, no spaces escaped (Acrobat tolerates raw)."""
    p = abs_path.replace("\\", "/")
    return f"file:///{p}#page={page}"


def read_csv(path: str) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_data_sheet(ws, rows: list[dict], headers: list[str], hyperlink_col: str | None,
                     hyperlink_target_fn=None, color_by_discipline=True):
    """Write rows to ws using the given header order. Optionally apply a hyperlink to one column."""
    HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
    HEADER_FONT = Font(bold=True)

    # Header
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")

    # Data
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            v = row.get(h, "")
            cell = ws.cell(row=i, column=j, value=v)
            if h == hyperlink_col and hyperlink_target_fn:
                target = hyperlink_target_fn(row)
                if target:
                    cell.value = f'=HYPERLINK("{target}","{v}")'
                    cell.font = Font(color="0563C1", underline="single")
            if color_by_discipline and h != hyperlink_col:
                disc = row.get("Discipline", "")
                color = DISCIPLINE_COLOR_MAP.get(disc)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

    # Freeze top row, autofilter, sized columns
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Auto-size by sampling content
    for j, h in enumerate(headers, start=1):
        sample = max((len(str(r.get(h, ""))) for r in rows[:200]), default=8)
        width = max(len(h), min(sample, 80)) + 2
        # Heuristic clamp per column meaning
        if h == "Page Title":
            width = max(width, 40)
        elif h in ("Sheet Number", "PDF Page"):
            width = max(width, 12)
        elif h == "Discipline":
            width = max(width, 22)
        elif h == "Source Issue":
            width = max(width, 28)
        ws.column_dimensions[get_column_letter(j)].width = min(width, 80)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("per_issue_csvs", nargs="+",
                    help="Per-issue Drawing Index CSVs in chronological order.")
    ap.add_argument("--franken", required=True, help="Path to the Franken Set CSV.")
    ap.add_argument("--anomaly", action="append", default=[],
                    help="Optional anomaly CSV (repeatable).")
    ap.add_argument("--out", default=None, help="Output xlsx path.")
    ap.add_argument("--search-root", action="append", default=[],
                    help="Extra folder to scan for source PDFs (repeatable).")
    ap.add_argument("--franken-pdf-folder", default=None,
                    help="Folder containing the Franken Set discipline PDFs (default: alongside Franken CSV in 'Drawings by Discipline/').")
    args = ap.parse_args()

    # Default search roots: walk up from the Franken Set CSV
    search_roots = list(args.search_root)
    franken_dir = os.path.dirname(os.path.abspath(args.franken))
    for candidate in [franken_dir, os.path.dirname(franken_dir), os.path.dirname(os.path.dirname(franken_dir))]:
        if candidate and os.path.isdir(candidate) and candidate not in search_roots:
            search_roots.append(candidate)
    pdf_index = build_pdf_index(search_roots)
    print(f"[info] indexed {len(pdf_index)} PDFs across {len(search_roots)} search roots", file=sys.stderr)

    # Where do the Franken Set discipline PDFs live?
    franken_pdf_folder = args.franken_pdf_folder or os.path.join(franken_dir, "Drawings by Discipline")
    franken_pdfs_available = os.path.isdir(franken_pdf_folder)
    if franken_pdfs_available:
        print(f"[info] Franken Set discipline PDFs at: {franken_pdf_folder}", file=sys.stderr)
    else:
        print(f"[warn] Franken Set discipline PDFs not found at {franken_pdf_folder} — hyperlinks fall back to source PDFs", file=sys.stderr)

    # ----- Load all CSVs -----
    per_issue: "OrderedDict[str, list[dict]]" = OrderedDict()
    for csv_path in args.per_issue_csvs:
        label = issue_label_from(csv_path)
        per_issue[label] = read_csv(csv_path)
        print(f"[info] {label}: {len(per_issue[label])} rows", file=sys.stderr)

    franken_rows = read_csv(args.franken)
    print(f"[info] Franken Set: {len(franken_rows)} rows", file=sys.stderr)

    anomaly_rows: list[dict] = []
    for ap_path in args.anomaly:
        for r in read_csv(ap_path):
            r["_source"] = os.path.basename(ap_path)
            anomaly_rows.append(r)
    if anomaly_rows:
        print(f"[info] Anomalies: {len(anomaly_rows)} rows", file=sys.stderr)

    # ----- Compute Franken Set new-page-index map (for hyperlinks into discipline PDFs) -----
    # For each discipline, sort its sheets and assign sequential 1-based page numbers.
    franken_disc_page: dict[tuple[str, str], int] = {}  # (discipline, sheet#) -> new_page
    by_disc: dict[str, list[dict]] = {}
    for r in franken_rows:
        by_disc.setdefault(r["Discipline"], []).append(r)
    for disc, rs in by_disc.items():
        rs.sort(key=lambda r: sheet_sort_key(r["Sheet Number"]))
        for i, r in enumerate(rs, start=1):
            franken_disc_page[(disc, r["Sheet Number"])] = i

    def per_issue_hyperlink(row):
        src = row.get("PDF File", "")
        page = row.get("Page in PDF", "")
        if not src or not page:
            return None
        abs_path = pdf_index.get(src)
        if not abs_path:
            return None
        try:
            return to_file_uri(abs_path, int(page))
        except ValueError:
            return None

    def franken_hyperlink(row):
        disc = row["Discipline"]
        sheet = row["Sheet Number"]
        if franken_pdfs_available:
            # Point at the new Franken discipline PDF + new page index
            fname = f"Franken Set -- {safe_filename(disc.title())}.pdf"
            fp = os.path.join(franken_pdf_folder, fname)
            if os.path.isfile(fp):
                new_page = franken_disc_page.get((disc, sheet))
                if new_page:
                    return to_file_uri(os.path.abspath(fp), new_page)
        # Fallback: link to the source PDF
        src = row.get("Source PDF File", "")
        page = row.get("Source Page in PDF", "")
        if src and page:
            abs_path = pdf_index.get(src)
            if abs_path:
                try:
                    return to_file_uri(abs_path, int(page))
                except ValueError:
                    return None
        return None

    # ----- Build the workbook -----
    wb = Workbook()
    wb.remove(wb.active)

    # Per-issue tabs
    for label, rows in per_issue.items():
        ws = wb.create_sheet(tab_name_from_issue(label))
        # Map CSV columns to display columns
        display_rows = []
        for r in rows:
            display_rows.append({
                "Discipline": r.get("Discipline", ""),
                "Sheet Number": r.get("Sheet Number", ""),
                "Page Title": r.get("Page Title", ""),
                "PDF Page": r.get("Page in PDF", ""),
                "Open": "Open",
                # carried for hyperlink fn
                "PDF File": r.get("PDF File", ""),
                "Page in PDF": r.get("Page in PDF", ""),
            })
        write_data_sheet(
            ws,
            display_rows,
            headers=["Discipline", "Sheet Number", "Page Title", "PDF Page", "Open"],
            hyperlink_col="Open",
            hyperlink_target_fn=per_issue_hyperlink,
        )

    # Franken Set tab (canonical)
    franken_tab = wb.create_sheet("Franken Set")
    franken_tab.sheet_properties.tabColor = "C00000"  # dark red — calls attention
    franken_display = []
    for r in franken_rows:
        franken_display.append({
            "Discipline": r.get("Discipline", ""),
            "Sheet Number": r.get("Sheet Number", ""),
            "Page Title": r.get("Page Title", ""),
            "Source Issue": r.get("Source Issue", ""),
            "Latest PDF Page": r.get("Source Page in PDF", ""),
            "Open": "Open",
            "Source PDF File": r.get("Source PDF File", ""),
            "Source Page in PDF": r.get("Source Page in PDF", ""),
        })
    # Sort by discipline then sheet # for the canonical view
    franken_display.sort(key=lambda r: (
        DISCIPLINE_ORDER.get(r["Discipline"], 999), sheet_sort_key(r["Sheet Number"]),
    ))
    write_data_sheet(
        franken_tab,
        franken_display,
        headers=["Discipline", "Sheet Number", "Page Title", "Source Issue", "Latest PDF Page", "Open"],
        hyperlink_col="Open",
        hyperlink_target_fn=franken_hyperlink,
    )

    # Anomalies tab
    if anomaly_rows:
        ws = wb.create_sheet("Anomalies")
        # Normalize columns across the different anomaly CSV schemas
        display = []
        for r in anomaly_rows:
            display.append({
                "Source": r.get("_source", ""),
                "Issue": r.get("PDF File", "") or "",
                "Sheet Number": r.get("Sheet Number", ""),
                "Discipline": r.get("Discipline", ""),
                "Type": r.get("Type", "unlisted"),
                "Page in PDF": r.get("Page in PDF", ""),
                "Note": r.get("Note", ""),
            })
        write_data_sheet(
            ws, display,
            headers=["Source", "Issue", "Sheet Number", "Discipline", "Type", "Page in PDF", "Note"],
            hyperlink_col=None,
            color_by_discipline=False,
        )
        ws.sheet_properties.tabColor = "ED7D31"

    # Reorder tabs: per-issue (chronological), then Franken Set, then Anomalies
    desired_order = [tab_name_from_issue(lbl) for lbl in per_issue.keys()] + ["Franken Set"]
    if anomaly_rows:
        desired_order.append("Anomalies")
    wb._sheets = sorted(wb._sheets, key=lambda s: desired_order.index(s.title) if s.title in desired_order else 999)

    # ----- Save (with Excel-lock versioning) -----
    out_path = args.out
    if not out_path:
        # Default: alongside the per-issue CSVs
        out_dir = os.path.dirname(franken_dir)
        out_path = os.path.join(out_dir, "Master Drawing Index.xlsx")

    target = out_path
    base, ext = os.path.splitext(out_path)
    n = 2
    while True:
        try:
            wb.save(target)
            break
        except PermissionError:
            target = f"{base} v{n}{ext}"
            n += 1
    print(f"[ok] wrote {target}")
    print(f"  tabs: {[s.title for s in wb._sheets]}", file=sys.stderr)


if __name__ == "__main__":
    main()
